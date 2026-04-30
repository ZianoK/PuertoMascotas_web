from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from io import BytesIO
import openpyxl
import re

from database import get_db
import models
from auth import verify_password, create_access_token, get_current_admin

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
)


# ---- Auth ----

class LoginRequest(BaseModel):
    email: str
    password: str


@router.post("/login")
def admin_login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.AdminUser).filter(
        models.AdminUser.email == req.email,
        models.AdminUser.active == True,
    ).first()

    if not user or not verify_password(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")

    token = create_access_token({"sub": user.email})
    return {"access_token": token, "token_type": "bearer", "name": user.name}


def _parse_price(value) -> float:
    """
    Convierte un precio argentino a float.
    Soporta: '$16.771,11', '16771,11', 16771.11, '$1.000', etc.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    s = s.replace("$", "").replace(" ", "").replace("\xa0", "")
    if "," in s and "." in s:
        # Formato argentino: punto=miles, coma=decimal → '16.771,11'
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        nums = re.findall(r"[\d\.]+", s)
        return float(nums[0]) if nums else 0.0
    except Exception:
        return 0.0


def _slugify_simple(text: str) -> str:
    """Slug simple sin dependencia externa."""
    import unicodedata
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text).strip().lower()
    return re.sub(r"[\s_-]+", "-", text)


@router.post("/products/upload")
async def upload_catalog(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents))

        created_categories: dict = {}
        products_created = 0
        products_updated = 0

        KNOWN_HEADERS = {
            "codigo": "codigo",
            "código": "codigo",
            "categoria": "categoria",
            "categoría": "categoria",
            "marca": "marca",
            "linea": "marca",
            "línea": "marca",
            "marca / línea": "marca",
            "marca / linea": "marca",
            "producto": "producto",
            "descripcion": "producto",
            "descripción": "producto",
            "precio": "precio",
            "precio ref.": "precio",
            "precio ref": "precio",
            "stock": "stock",
            "cantidad": "stock",
            "cant": "stock",
            "cant.": "stock",
        }

        for sheet in wb.worksheets:
            header_row_idx = None
            col_map: dict = {}

            # Detectar fila de encabezados (máx primeras 20 filas)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
                if not row:
                    continue
                row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
                matches = sum(1 for cell in row_lower if cell in KNOWN_HEADERS)
                if matches >= 2:
                    header_row_idx = row_idx + 1
                    for i, cell in enumerate(row_lower):
                        if cell in KNOWN_HEADERS:
                            col_map[KNOWN_HEADERS[cell]] = i
                    break

            data_start = (header_row_idx + 1) if header_row_idx else 2
            use_direct_map = bool(col_map)

            for row in sheet.iter_rows(min_row=data_start, values_only=True):
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue

                try:
                    if use_direct_map:
                        def get_col(key: str, default=None):
                            idx = col_map.get(key)
                            if idx is not None and idx < len(row):
                                v = row[idx]
                                return str(v).strip() if v is not None else default
                            return default

                        code = get_col("codigo") or ""
                        cat_name = get_col("categoria", "General") or "General"
                        marca = get_col("marca", "") or ""
                        producto = get_col("producto", "") or ""

                        if marca and producto:
                            name = f"{marca} {producto}"
                        elif producto:
                            name = producto
                        elif marca:
                            name = marca
                        else:
                            continue

                        price_raw = None
                        if "precio" in col_map and col_map["precio"] < len(row):
                            price_raw = row[col_map["precio"]]
                        price = _parse_price(price_raw)

                        stock = 0
                        if "stock" in col_map and col_map["stock"] < len(row):
                            try:
                                sv = row[col_map["stock"]]
                                if sv is not None:
                                    stock = int(float(str(sv).replace(",", ".")))
                            except (ValueError, TypeError):
                                pass

                    else:
                        code = ""
                        texts = []
                        price = 0.0
                        stock = 0
                        for cell in row:
                            s = str(cell).strip() if cell is not None else ""
                            if not s:
                                continue
                            if "$" in s or re.match(r"^\d[\d\.,]+$", s):
                                p = _parse_price(s)
                                if p > 100 and price == 0.0:
                                    price = p
                            else:
                                try:
                                    v = float(s.replace(",", "."))
                                    if v < 10000 and stock == 0:
                                        stock = int(v)
                                except ValueError:
                                    if len(s) > 1:
                                        texts.append(s)

                        if not texts:
                            continue
                        cat_name = "General"
                        name = " ".join(texts[:3])

                    name = name[:200].strip()
                    if not name:
                        continue
                    cat_name = (cat_name or "General")[:50].strip()

                    cat_slug = _slugify_simple(cat_name)[:50] or "general"

                    if cat_slug not in created_categories:
                        db_cat = db.query(models.Category).filter(models.Category.slug == cat_slug).first()
                        if not db_cat:
                            db_cat = models.Category(name=cat_name, slug=cat_slug)
                            db.add(db_cat)
                            db.flush()
                        created_categories[cat_slug] = db_cat.id

                    cat_id = created_categories[cat_slug]

                    db_product = db.query(models.Product).filter(models.Product.name == name[:200]).first()
                    if db_product:
                        db_product.price = price
                        if stock > 0:
                            db_product.stock = stock
                        if code:
                            db_product.code = code[:50]
                        db_product.category_id = cat_id
                        products_updated += 1
                    else:
                        db_product = models.Product(
                            name=name[:200],
                            code=code[:50] if code else None,
                            description=f"Importado desde catálogo. Categoría: {cat_name}",
                            price=price,
                            cost=0.0,
                            stock=stock,
                            category_id=cat_id,
                            image_url=""
                        )
                        db.add(db_product)
                        products_created += 1

                except Exception as e:
                    print(f"[Excel] Error en fila {row}: {e}")
                    continue

        if products_created == 0 and products_updated == 0:
            raise HTTPException(
                status_code=400,
                detail="No se encontraron productos. Verificá que el Excel tenga columnas de Nombre/Producto y Precio."
            )

        db.commit()
        return {
            "message": "Catálogo procesado correctamente",
            "products_created": products_created,
            "products_updated": products_updated
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products")
def get_admin_products(db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    products = db.query(models.Product).all()
    result = []
    for p in products:
        result.append({
            "id": p.id,
            "code": p.code or "",
            "name": p.name,
            "price": p.price,
            "original_price": p.original_price,
            "stock": p.stock,
            "category": p.category.name if p.category else "General"
        })
    return result


@router.delete("/products/{id}")
def delete_product(id: int, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    product = db.query(models.Product).filter(models.Product.id == id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(product)
    db.commit()
    return {"message": "Product deleted"}


class StockRequest(BaseModel):
    stock: int


class DiscountRequest(BaseModel):
    new_price: float


@router.put("/products/{id}/stock")
def update_stock(id: int, req: StockRequest, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    product = db.query(models.Product).filter(models.Product.id == id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    product.stock = req.stock
    db.commit()
    return {"message": "Stock updated", "stock": product.stock}


@router.put("/products/{id}/discount")
def apply_discount(id: int, req: DiscountRequest, db: Session = Depends(get_db), admin: models.AdminUser = Depends(get_current_admin)):
    product = db.query(models.Product).filter(models.Product.id == id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if req.new_price <= 0:
        raise HTTPException(status_code=400, detail="El precio debe ser mayor a 0")
    if req.new_price >= product.price:
        raise HTTPException(status_code=400, detail="El precio con descuento debe ser menor al precio actual")

    if not product.original_price:
        product.original_price = product.price

    product.price = req.new_price
    db.commit()
    return {"message": "Discount applied", "price": product.price, "original_price": product.original_price}
